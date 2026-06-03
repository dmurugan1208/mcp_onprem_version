"""
REQ-09: Generic Document Retrieval — BM25 full-text search across all worker files.

One document = one BM25 entry (no chunking).

Cache invalidation: fingerprint-based — fast directory scan (file paths + mtimes) on every
search call. If the fingerprint matches the cached snapshot, reuse the index. If anything
changed (new file uploaded, file modified/deleted), rebuild immediately. No TTL wait.

Supported file types: .md .txt .py .rst .yaml .yml .json .csv .docx .pdf .xlsx
"""

import io
import os
import logging
from typing import Optional

from rank_bm25 import BM25Okapi

from sajha.tools.base_mcp_tool import BaseMCPTool
from sajha.core.properties_configurator import PropertiesConfigurator
from sajha.storage import storage

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {
    '.md', '.txt', '.py', '.rst',
    '.yaml', '.yml', '.json', '.csv',
    '.docx', '.pdf', '.xlsx',
}

# ---------------------------------------------------------------------------
# Module-level cache
# ---------------------------------------------------------------------------
# key  : "<domain_dir>|<my_data_dir>"
# value: {"fingerprint": dict[path→mtime], "bm25": BM25Okapi|None, "docs": list}
_INDEX_CACHE: dict = {}


# ---------------------------------------------------------------------------
# Fingerprint — fast freshness check, no file reading
# ---------------------------------------------------------------------------

def _fingerprint(directories: list) -> dict:
    """Return {abs_path: mtime} for every file under the given directories."""
    fp = {}
    for d in directories:
        if not storage.exists(d):
            continue
        for rel in storage.list_prefix(d):
            abs_path = os.path.join(d, rel)
            fp[abs_path] = storage.get_size(abs_path)
    return fp


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------

def _extract_text(abs_path: str) -> Optional[str]:
    """Return plain text for a file, or None if unsupported / unreadable."""
    ext = os.path.splitext(abs_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        return None
    try:
        raw = storage.read_bytes(abs_path)
    except Exception as exc:
        logger.debug("BM25: cannot read %s: %s", abs_path, exc)
        return None

    try:
        if ext in ('.md', '.txt', '.py', '.rst', '.yaml', '.yml', '.csv', '.json'):
            return raw.decode('utf-8', errors='replace')

        if ext == '.docx':
            from docx import Document as _Document
            doc = _Document(io.BytesIO(raw))
            return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())

        if ext == '.pdf':
            import pdfplumber
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                return '\n'.join(page.extract_text() or '' for page in pdf.pages)

        if ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cell_text = ' '.join(str(c) for c in row if c is not None)
                    if cell_text.strip():
                        rows.append(cell_text)
            return '\n'.join(rows)

    except Exception as exc:
        logger.debug("BM25: extraction failed for %s: %s", abs_path, exc)

    return None


def _extract_excerpt(text: str, query_tokens: list, context_chars: int = 350) -> str:
    """Return a short excerpt around the first matching query token."""
    lower = text.lower()
    for token in query_tokens:
        pos = lower.find(token)
        if pos >= 0:
            half = context_chars // 2
            start = max(0, pos - half)
            end = min(len(text), pos + len(token) + half)
            excerpt = text[start:end].strip()
            return ('...' if start > 0 else '') + excerpt + ('...' if end < len(text) else '')
    return text[:context_chars].strip() + ('...' if len(text) > context_chars else '')


# ---------------------------------------------------------------------------
# Index builder — one doc per file, no chunking
# ---------------------------------------------------------------------------

def _build_index(directories: list) -> tuple:
    """
    Scan directories, extract text, build BM25Okapi index.
    One document per file — no chunking.
    Returns (bm25, docs_list) — bm25 is None if no indexable content found.
    """
    docs = []
    seen = set()

    for directory in directories:
        if not storage.exists(directory):
            continue
        for rel in storage.list_prefix(directory):
            abs_path = os.path.join(directory, rel)
            if abs_path in seen:
                continue
            seen.add(abs_path)
            text = _extract_text(abs_path)
            if not text or not text.strip():
                continue
            docs.append({
                'file_path': abs_path,
                'file_name': os.path.basename(abs_path),
                'text': text.strip(),
            })

    if not docs:
        return None, docs

    tokenized = [d['text'].lower().split() for d in docs]
    return BM25Okapi(tokenized), docs


# ---------------------------------------------------------------------------
# Tool
# ---------------------------------------------------------------------------

class DocumentSearchTool(BaseMCPTool):
    """
    BM25 full-text search across domain_data and my_data files.
    One entry per document. Fingerprint-based cache — automatically picks up
    newly uploaded files on the next search call with no wait.
    """

    def get_input_schema(self) -> dict:
        return self.config.get('inputSchema', {
            'type': 'object',
            'properties': {'query': {'type': 'string'}},
            'required': ['query'],
        })

    def get_output_schema(self) -> dict:
        return {'type': 'object'}

    def _domain_dir(self) -> str:
        try:
            from flask import g as _g
            root = getattr(_g, 'worker_data_root', None)
            if root:
                return root.rstrip('/')
        except RuntimeError:
            pass
        prop = PropertiesConfigurator().get('data.domain_data.dir', '')
        if prop:
            return prop
        raise RuntimeError(
            'BM25Search: no worker context available and data.domain_data.dir '
            'is not set in application.properties'
        )

    def _my_data_dir(self) -> str:
        try:
            from flask import g as _g
            root = getattr(_g, 'worker_my_data_root', None)
            if root:
                return root.rstrip('/')
        except RuntimeError:
            pass
        prop = PropertiesConfigurator().get('data.my_data.dir', '')
        if prop:
            return prop
        raise RuntimeError(
            'BM25Search: no worker context available and data.my_data.dir '
            'is not set in application.properties'
        )

    def _common_dir(self) -> str:
        try:
            from flask import g as _g
            root = getattr(_g, 'worker_common_root', None)
            if root:
                return root.rstrip('/')
        except RuntimeError:
            pass
        return PropertiesConfigurator().get(
            'data.common_data.dir',
            './data/common',
        )

    def execute(self, arguments: dict) -> dict:
        query = str(arguments.get('query', '')).strip()
        if not query:
            return {'error': 'query is required', 'results': []}

        top_k = max(1, int(arguments.get('top_k', 10)))
        top_n_full = max(0, int(arguments.get('top_n_full_content', 0)))
        file_type_filter = [
            ft.lower() if ft.startswith('.') else '.' + ft.lower()
            for ft in (arguments.get('file_types') or [])
        ]

        domain_dir = self._domain_dir()
        my_data_dir = self._my_data_dir()
        common_dir = self._common_dir()
        cache_key = f"{domain_dir}|{my_data_dir}|{common_dir}"

        # ── Fingerprint check — rebuild only when files changed ──────
        current_fp = _fingerprint([domain_dir, my_data_dir, common_dir])
        cached = _INDEX_CACHE.get(cache_key)
        rebuilt = False

        if cached is None or cached['fingerprint'] != current_fp:
            bm25, docs = _build_index([domain_dir, my_data_dir, common_dir])
            _INDEX_CACHE[cache_key] = {
                'fingerprint': current_fp,
                'bm25': bm25,
                'docs': docs,
            }
            rebuilt = True
        else:
            bm25 = cached['bm25']
            docs = cached['docs']

        if bm25 is None or not docs:
            return {
                'query': query,
                'total_results': 0,
                'index_size': 0,
                'rebuilt': rebuilt,
                'results': [],
                'message': 'No indexable documents found in domain_data or my_data.',
            }

        # ── Optional file-type filter ─────────────────────────────────
        if file_type_filter:
            filtered_idx = [
                i for i, d in enumerate(docs)
                if os.path.splitext(d['file_name'])[1].lower() in file_type_filter
            ]
        else:
            filtered_idx = list(range(len(docs)))

        # ── BM25 scoring ──────────────────────────────────────────────
        query_tokens = query.lower().split()
        all_scores = bm25.get_scores(query_tokens)

        scored = sorted(
            ((i, float(all_scores[i])) for i in filtered_idx if all_scores[i] > 0),
            key=lambda x: x[1],
            reverse=True,
        )[:top_k]

        # ── Build results ─────────────────────────────────────────────
        results = []
        for rank, (idx, score) in enumerate(scored):
            doc = docs[idx]
            result = {
                'rank': rank + 1,
                'file_name': doc['file_name'],
                'file_path': doc['file_path'],
                'score': round(score, 4),
                'excerpt': _extract_excerpt(doc['text'], query_tokens),
            }
            if rank < top_n_full:
                result['full_content'] = doc['text']
            results.append(result)

        return {
            'query': query,
            'total_results': len(results),
            'index_size': len(docs),
            'rebuilt': rebuilt,
            'results': results,
        }
